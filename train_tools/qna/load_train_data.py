import mysql.connector
import openpyxl
import os
from chatbot.config.DatabaseConfig import *

# # from chatbot.utils.utils import get_abs_path

# # train_file2 = get_abs_path('train_data.xlsx')
# # wb2 = openpyxl.load_workbook(train_file2)

# print('현재 작업 디렉토리')
# print(os.getcwd())

def all_clear_train_data(cnx):
    sql = """
        DELETE FROM chatbot_train_data
        """
    
    with cnx.cursor() as cursor:
        cursor.execute(sql)

    sql = """
        ALTER TABLE chatbot_train_data AUTO_INCREMENT=1
        """
    with cnx.cursor() as cursor:
        cursor.execute(sql)

def insert_data(cnx, xls_row):
    intent, ner, query, answer, answer_img_url = xls_row

    sql = f"""
        INSERT chatbot_train_data(intent, ner, query, answer, answer_image)
        values(
        '{intent.value}','{ner.value}','{query.value}','{answer.value}','{answer_img_url.value}'
        )
    """

    sql = sql.replace("'None'", "null")

    with cnx.cursor() as cursor:
        cursor.execute(sql)
        print(f"{query.value} 저장")
        cnx.commit()

train_file = 'project/chatbot/train_tools/qna/train_data.xlsx'
cnx = None

try :
    cnx = mysql.connector.connect(user=DB_USER, password=DB_PASSWORD, 
                                host=DB_HOST, port=3306, database=DB_NAME, ssl_ca="{ca-cert filename}", 
                                ssl_disabled=False)
    
    all_clear_train_data(cnx)

    wb = openpyxl.load_workbook(train_file)
    sheet = wb['Sheet1']
    for row in sheet.iter_rows(min_row=2):
        insert_data(cnx, row)

    wb.close()

except Exception as e:
    print(e)

finally:
    if cnx is not None:
        cnx.close()